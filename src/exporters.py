from __future__ import annotations

from io import BytesIO
from typing import Dict, Optional
import pandas as pd
import numbers
import re
import unicodedata
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

def _norm(s: str) -> str:
    """Uppercase, remove accents, normalize spaces."""
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def format_ar_number(value: float, decimals: int = 2) -> str:
    """Formatea número con miles '.' y decimales ',' (Argentina)."""
    try:
        v = float(value)
    except Exception:
        return str(value)
    s = f"{v:,.{decimals}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

def _col_kind(colname: str) -> str:
    up = (colname or "").upper()
    if re.search(r"(NETO|IVA|GASTO|IMPORTE|MONTO|BRUTO|TOTAL|PRECIO|COMISI|OTROS)", up):
        return "money"
    if "KILO" in up:
        return "kilos"
    if re.search(r"(CABEZA|CANTIDAD)", up):
        return "qty"
    return "text"


def df_to_pdf_bytes(df: pd.DataFrame, title: str = "Grilla") -> bytes:
    """
    Exporta un DataFrame a PDF (tabla) con layout estable (A4 horizontal).
    - Evita encabezados "rotos" (no usa Paragraph en header).
    - Para Grillas de Compras/Ventas usa un set compacto de columnas para que entren prolijas.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 10)]

    if df is None or df.empty:
        story.append(Paragraph("Sin datos.", styles["Normal"]))
        doc.build(story)
        return buffer.getvalue()

    # --- Selección de columnas para PDF (para que NO quede horrible) ---
    cols = list(df.columns)

    def _pick(names: list[str]) -> list[str]:
        picked = []
        for n in names:
            if n in cols:
                picked.append(n)
        return picked

    title_norm = _norm(title)
    is_grid = ("GRILLA" in title_norm)

    if is_grid:
        # Orden compacto (mantener similar a tu compras.pdf "bueno" + Categoria/Raza)
        desired = [
            "Fecha",
            "Fecha Operación",
            "Título",
            "Tipo",
            "Cód ARCA",
            "Letra",
            "PV",
            "Número",
            "Ajuste",
            "Ajuste sentido",
            "Ajuste tipo",
            "Contraparte CUIT",
            "Contraparte",
            "Cond IVA",
            "Categoría/Raza",
            "Cabezas",
            "Kilos",
            "Neto Hacienda (sin gastos)",
            "IVA Hacienda",
            "Gastos (sin IVA)",
            "IVA Gastos",
        ]
        sel = _pick(desired)
        # Si faltan algunas (según reporte), igual seguimos con lo que haya
        if sel:
            df_pdf = df[sel].copy()
        else:
            df_pdf = df.copy()
    else:
        df_pdf = df.copy()

    # --- Preparar datos ---
    # Header como string (para que NO rompa palabras en vertical)
    header = [str(c) for c in df_pdf.columns]

    # Columnas texto que conviene wrappear
    text_cols = set()
    for idx, c in enumerate(df_pdf.columns):
        c_norm = _norm(str(c))
        if any(k in c_norm for k in ("CATEG", "RAZA", "CONTRAPARTE", "RAZON", "TITULO")):
            text_cols.add(idx)

    kinds = [_col_kind(c) for c in df_pdf.columns]

    # Estilo compacto para celdas wrap
    body_style = styles["BodyText"].clone("body_small")
    body_style.fontSize = 6.5
    body_style.leading = 7.5

    data = [header]
    for _, row in df_pdf.iterrows():
        out_row = []
        for col_idx, (v, k) in enumerate(zip(row.values, kinds)):
            if pd.isna(v):
                out_row.append("")
                continue

            if isinstance(v, numbers.Number) and k in ("money", "kilos"):
                out_row.append(format_ar_number(v, decimals=2))
            elif isinstance(v, numbers.Number) and k == "qty":
                out_row.append(format_ar_number(v, decimals=0))
            else:
                s = str(v)
                if col_idx in text_cols and len(s) > 28:
                    out_row.append(Paragraph(s, body_style))
                else:
                    out_row.append(s)
        data.append(out_row)

    # --- Anchos de columnas (layout estable) ---
    page_w, _ = landscape(A4)
    avail_w = page_w - doc.leftMargin - doc.rightMargin

    base_widths = []
    for c, k in zip(df_pdf.columns, kinds):
        cn = _norm(str(c))
        if "CATEG" in cn or "RAZA" in cn:
            w = 260
        elif "CONTRAPARTE" in cn and "CUIT" not in cn:
            w = 150
        elif "CONTRAPARTE" in cn and "CUIT" in cn:
            w = 85
        elif "TITULO" in cn:
            w = 120
        elif "FECHA" in cn:
            w = 70
        elif "NUMERO" in cn:
            w = 70
        elif "PV" in cn:
            w = 45
        elif k == "qty":
            w = 45
        elif k in ("money", "kilos"):
            w = 78
        elif "COD" in cn:
            w = 55
        elif "LETRA" in cn:
            w = 35
        elif "UM" in cn:
            w = 45
        else:
            w = 60
        base_widths.append(w)

    total_w = sum(base_widths)
    if total_w > avail_w:
        scale = avail_w / total_w
        col_widths = [max(32, w * scale) for w in base_widths]
    else:
        col_widths = base_widths

    table = Table(data, repeatRows=1, colWidths=col_widths)

    style = TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ])

    # Alinear números a derecha
    for j, k in enumerate(kinds):
        if k in ("money", "kilos", "qty"):
            style.add("ALIGN", (j, 1), (j, -1), "RIGHT")

    table.setStyle(style)
    story.append(table)
    doc.build(story)
    return buffer.getvalue()

    # Preparar datos (strings) y wrapping en columnas texto largas
    kinds = [_col_kind(c) for c in df.columns]
    header = [Paragraph(str(c), styles["BodyText"]) for c in df.columns]
    data = [header]

    text_cols = set()
    for idx, c in enumerate(df.columns):
        c_norm = _norm(str(c))
        if any(k in c_norm for k in ("CATEG", "RAZA", "CONTRAPARTE", "RAZON", "TITULO")):
            text_cols.add(idx)

    for _, row in df.iterrows():
        out_row = []
        for col_idx, (v, k) in enumerate(zip(row.values, kinds)):
            if pd.isna(v):
                out_row.append("")
                continue
            if isinstance(v, numbers.Number) and k in ("money", "kilos"):
                out_row.append(format_ar_number(v, decimals=2))
            elif isinstance(v, numbers.Number) and k == "qty":
                out_row.append(format_ar_number(v, decimals=0))
            else:
                s = str(v)
                # Wrap en columnas texto
                if col_idx in text_cols and len(s) > 40:
                    out_row.append(Paragraph(s, styles["BodyText"]))
                else:
                    out_row.append(s)
        data.append(out_row)

    # Anchos: asignar por tipo de columna y escalar a página
    page_w, _ = landscape(A4)
    avail_w = page_w - doc.leftMargin - doc.rightMargin

    base_widths = []
    for c, k in zip(df.columns, kinds):
        cn = _norm(str(c))
        if "CATEG" in cn or "RAZA" in cn:
            w = 260
        elif "CONTRAPARTE" in cn or "RAZON" in cn:
            w = 190
        elif k == "qty":
            w = 55
        elif k in ("money", "kilos"):
            w = 80
        elif "FECHA" in cn:
            w = 70
        elif "NUMERO" in cn:
            w = 80
        elif "PV" in cn or "PUNTO" in cn:
            w = 55
        elif "UM" in cn:
            w = 55
        else:
            w = 65
        base_widths.append(w)

    total_w = sum(base_widths)
    if total_w > avail_w:
        scale = avail_w / total_w
        col_widths = [max(35, w * scale) for w in base_widths]
    else:
        col_widths = base_widths

    table = Table(data, repeatRows=1, colWidths=col_widths)
    style = TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ])
    table.setStyle(style)
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def dfs_to_excel_bytes(sheets: Dict[str, pd.DataFrame]) -> bytes:
    """
    Exporta múltiples DataFrames a un solo Excel (xlsx) y aplica formato numérico.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = (name or "Hoja")[:31]
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=safe_name, index=False)

            ws = writer.book[safe_name]
            # Aplicar formatos por columna (monto/cantidad/kilos)
            for j, col in enumerate(df.columns, start=1):
                kind = _col_kind(str(col))
                if kind == "money":
                    fmt = '#,##0.00'
                elif kind == "kilos":
                    fmt = '#,##0.00'
                elif kind == "qty":
                    fmt = '#,##0'
                else:
                    fmt = None
                if fmt:
                    for i in range(2, ws.max_row + 1):
                        cell = ws.cell(i, j)
                        if isinstance(cell.value, numbers.Number):
                            # Cantidades: forzar entero visible
                            if kind == 'qty':
                                try:
                                    cell.value = int(round(float(cell.value)))
                                except Exception:
                                    pass
                            cell.number_format = fmt
            # Ajuste ancho columnas (según contenido)
            for j, col in enumerate(df.columns, start=1):
                letter = get_column_letter(j)
                try:
                    max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values])
                except Exception:
                    max_len = len(str(col))
                width = min(80, max(12, max_len + 2))
                # Prioridad: Tipo de Hacienda suele ser largo
                if "HACIENDA" in str(col).upper():
                    width = max(width, 55)
                ws.column_dimensions[letter].width = width
    return output.getvalue()



from openpyxl import load_workbook

def df_to_template_excel_bytes(template_path: str, df: pd.DataFrame, sheet_name: str = "Salida") -> bytes:
    """Carga un template XLSX y vuelca el DataFrame respetando encabezados del template.
    - Soporta columna interna `__bold__` (no se escribe; solo aplica negrita en la fila).
    - Aplica formatos numéricos compatibles con Excel (miles/decimales según configuración regional del usuario).
    """
    from openpyxl.styles import Font

    output = BytesIO()
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header_row = 1
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    col_to_idx = {h: i + 1 for i, h in enumerate(headers) if h not in (None, "")}

    # Limpiar filas existentes debajo del header
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    if df is None:
        df = pd.DataFrame()

    # helper: set number formats
    def _norm_header(name: str) -> str:
        s = (name or "")
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        return s.upper().strip()

    def _apply_fmt(cell, colname: str, val):
        # Normalizar header para evitar problemas con acentos: "Alíc." -> "ALIC."
        up = _norm_header(colname)

        if val is None or val == "":
            return

        # Prioridad: TD / Tipo Doc
        if up == "TD" or "TIPO DOC" in up:
            cell.number_format = "0"
            return

        # Prioridad: códigos (ej. "Cód. Neto", "Cód. NG/EX") -> entero sin decimales
        if "COD" in up or "CÓD" in up:
            cell.number_format = "0"
            # Forzar entero si viene como float (pandas)
            if isinstance(val, float) and val.is_integer():
                cell.value = int(val)
            return

        # Concepto (ventas) -> entero (141)
        if up == "CONCEPTO":
            cell.number_format = "0"
            if isinstance(val, float) and val.is_integer():
                cell.value = int(val)
            return

        if isinstance(val, (int, float)):
            # Alícuota con 3 decimales (10,500 / 21,000 / 0.000)
            if "ALIC" in up:
                cell.number_format = "0.000"
                return

            # Cantidades (cabezas / unidades) sin decimales
            if re.search(r"(CABEZA|CANTIDAD)", up):
                cell.number_format = "#,##0"
                return

            # Kilos / precios / importes: 2 decimales
            if "KILO" in up:
                cell.number_format = "#,##0.00"
                return

            if re.search(r"(NETO|IVA|GASTO|IMPORTE|MONTO|BRUTO|TOTAL|PRECIO|COMISI|OTROS|CONCEP)", up):
                cell.number_format = "#,##0.00"
                return

    # Escribir filas
    for r, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        is_bold = False
        if "__bold__" in row.index and bool(row.get("__bold__")):
            is_bold = True

        for col, val in row.items():
            if col == "__bold__":
                continue
            if col not in col_to_idx:
                continue
            c = col_to_idx[col]
            cell = ws.cell(r, c)
            if pd.isna(val):
                cell.value = None
            else:
                cell.value = val
            _apply_fmt(cell, col, cell.value)

        if is_bold:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).font = Font(bold=True)

    # Ancho de columnas (auto simple)
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        hdr = ws.cell(header_row, c).value
        if hdr:
            ws.column_dimensions[letter].width = min(45, max(10, len(str(hdr)) + 2))

    wb.save(output)
    return output.getvalue()
